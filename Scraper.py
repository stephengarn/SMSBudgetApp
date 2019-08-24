# libraries
from twilio.rest import Client
from selenium import webdriver
from openpyxl import Workbook, load_workbook
from selenium.webdriver.firefox.options import Options
# from selenium.webdriver.common.keys import Keys
import re
import time
import os
# variable and list declarations


#username and password for your financial institution
user = os.environ['MBNAUser']
passwd = os.environ['MBNAPass']
account = os.environ['twilioAccount']
auth = os.environ['twilioAuth']

dollarValues = []
digits = []
cells = []
sendText = False

# start the engines of webdriver and load banking site

driver = webdriver.Firefox()
options = Options()
options.add_argument('-headless')
driver.get(
	"https://authentication.mbna.ca/uap-ui/index.html?consumer=msec&locale=en_CA#/login/login/mbna/login/mbna/mbna-login")

# we are sleeping because we need to let the page load before we can start looking for elements

time.sleep(10)
def userAndPass(userOrPass, elemToFind):
	elem = driver.find_element_by_id(elemToFind)
	elem.send_keys(userOrPass)

def clickAndSleep(elem):
	elem.click()

userAndPass(user, "username100")
userAndPass(passwd, "password")

elem = driver.find_element_by_class_name("td-button")

clickAndSleep(elem)
time.sleep(8)


elem = driver.find_element_by_class_name("card-name")

clickAndSleep(elem)

table = driver.find_elements_by_id("transactionTable")
time.sleep(4)

# find the data

rows = table[0].find_elements_by_tag_name("td")
time.sleep(4)

# grab the data we need

for row in rows:
	x = re.findall(r"\$[^\]]+", row.text)
	# print(x)
	dollarValues.append(x)

# remove empty elements from list

dollarValues = list(filter(None, dollarValues))

# we only want the first five elemnts
for i in dollarValues[0:5]:
	i = str(i)
	i = i.replace("$", '')
	digits.append(i)

for i in range(len(digits)):
	digits[i] = digits[i].replace("['", "")

for i in range(len(digits)):
	digits[i] = digits[i].replace("']", "")

for i in range(len(digits)):
	digits[i] = float(digits[i])

# load in the excel stuff
wb = Workbook()
wb = load_workbook(filename='balances.xlsx')
ws = wb.active
balanceA10 = ws['A10']
for i in range(1, 6):
	i = str(i)
	appendValue = 'A' + i
	cells.append(appendValue)

# now we need to compare the digits with the cells values. If they do not match, we need to place the digits into the
# cell values.

cellsValues = []
for i in range(0, 5):
	cellsValues.append(ws[cells[i]])

cellValues = []
for i in range(0,5):
	cellValues.append(cellsValues[i].value)

#while(digits[i] != cellsValues[i].value):
#	cellsValues[i].value = digits[i]
#	balanceA10.value -= digits[i]
#	i += 1

for i in range(0, 5):
	if sortedDigits[i] != sortedCellValues[i]:
		cellsValues[i].value = sortedDigits[i]
		# let's deduct new amounts from
		#print(sortedDigits[i])
		balanceA10.value -= sortedDigits[i]
		sendText = True
	else:
		print('No new transactions.')

#we shall move onto the SMS part. We will text and give an update to ourselves about the remaining funds in our budget
if sendText:
	client = Client(account, auth)
	message = client.messages.create(
		body="You just spent " + str(digits[0]) + " and your remaining balance is " + str(round(balanceA10.value)),
		from_='+',
		to='+'
	)

wb.save('balances.xlsx')